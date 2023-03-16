from PyQt5 import QtCore, QtGui, QtWidgets
import os
from datetime import datetime
from PyQt5 import QtCore, QtGui, QtWidgets
import sqlite3
import openpyxl
baglanti = sqlite3.connect("./data.db")
cunsor = baglanti.cursor()
cunsor.execute("CREATE TABLE IF NOT EXISTS ana (isadi TEXT , firma TEXT , avansialan TEXT , verilen int )")
cunsor.execute("CREATE TABLE IF NOT EXISTS FİRMA (id İNT ,tarih TEXT , belge TEXT , FİRMA TEXT , Kişi int , gider TEXT ,toplam int )")
baglanti.commit()
import firebase_admin
from firebase_admin import credentials
from firebase_admin import firestore
class Ui_MainWindow(object):
    def setupUi(self, MainWindow):
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1253, 373)
        icon = QtGui.QIcon()
        icon.addPixmap(QtGui.QPixmap(":/icon/excel.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        MainWindow.setWindowIcon(icon)
        MainWindow.setStyleSheet("#centralwidget{\n"
"background-color: rgb(172, 255, 200);}")
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.widget = QtWidgets.QWidget(self.centralwidget)
        self.widget.setGeometry(QtCore.QRect(20, 40, 581, 281))
        self.widget.setObjectName("widget")
        self.pushButton = QtWidgets.QPushButton(self.widget)
        self.pushButton.setGeometry(QtCore.QRect(300, 119, 87, 43))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton.setFont(font)
        icon1 = QtGui.QIcon()
        icon1.addPixmap(QtGui.QPixmap(":/icon/note.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton.setIcon(icon1)
        self.pushButton.setIconSize(QtCore.QSize(22, 23))
        self.pushButton.setObjectName("pushButton")
        self.pushButton.clicked.connect(self.ileri)
        self.lineEdit = QtWidgets.QLineEdit(self.widget)
        self.lineEdit.setGeometry(QtCore.QRect(10, 120, 291, 41))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit.setFont(font)
        self.lineEdit.setText("")
        self.lineEdit.setObjectName("lineEdit")
        self.label = QtWidgets.QLabel(self.widget)
        self.label.setGeometry(QtCore.QRect(20, 60, 292, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.pushButton_3 = QtWidgets.QPushButton(self.widget)
        self.pushButton_3.setGeometry(QtCore.QRect(470, 150, 111, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_3.setFont(font)
        self.pushButton_3.setIcon(icon)
        self.pushButton_3.setIconSize(QtCore.QSize(24, 23))
        self.pushButton_3.setObjectName("pushButton_3")
        self.pushButton_3.clicked.connect(self.kaydet)
        self.label_2 = QtWidgets.QLabel(self.widget)
        self.label_2.setGeometry(QtCore.QRect(0, 0, 581, 21))
        self.label_2.setStyleSheet("#label_2{\n"
"\n"
"\n"
"background-image: url(:/icon/MERT.PNG);}")
        self.label_2.setText("")
        self.label_2.setObjectName("label_2")
        self.label_3 = QtWidgets.QLabel(self.widget)
        self.label_3.setGeometry(QtCore.QRect(0, 30, 581, 21))
        font = QtGui.QFont()
        font.setPointSize(13)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.label_6 = QtWidgets.QLabel(self.widget)
        self.label_6.setGeometry(QtCore.QRect(0, 230, 292, 51))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_6.setFont(font)
        self.label_6.setText("")
        self.label_6.setObjectName("label_6")
        self.label_7 = QtWidgets.QLabel(self.widget)
        self.label_7.setGeometry(QtCore.QRect(10, 160, 371, 41))
        font = QtGui.QFont()
        font.setPointSize(12)
        font.setBold(True)
        font.setWeight(75)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.widget)
        self.lineEdit_2.setGeometry(QtCore.QRect(10, 210, 291, 41))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.lineEdit_2.setFont(font)
        self.lineEdit_2.setText("")
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.pushButton_4 = QtWidgets.QPushButton(self.widget)
        self.pushButton_4.setGeometry(QtCore.QRect(300, 209, 87, 43))
        font = QtGui.QFont()
        font.setPointSize(11)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_4.setFont(font)
        self.pushButton_4.setIcon(icon)
        self.pushButton_4.setIconSize(QtCore.QSize(24, 23))
        self.pushButton_4.setObjectName("pushButton_4")
        self.pushButton_4.clicked.connect(self.mobil)
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(500, 300, 101, 20))
        font = QtGui.QFont()
        font.setPointSize(7)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.tableWidget = QtWidgets.QTableWidget(self.centralwidget)
        self.tableWidget.setGeometry(QtCore.QRect(620, 30, 591, 281))
        self.tableWidget.setRowCount(100)
        self.tableWidget.setObjectName("tableWidget")
        self.tableWidget.setColumnCount(6)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(8)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(0, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(7)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(1, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(7)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(2, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(7)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(3, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(7)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(4, item)
        item = QtWidgets.QTableWidgetItem()
        font = QtGui.QFont()
        font.setPointSize(7)
        item.setFont(font)
        self.tableWidget.setHorizontalHeaderItem(5, item)
        self.tableWidget.horizontalHeader().setDefaultSectionSize(88)
        self.tableWidget.horizontalHeader().setMinimumSectionSize(49)
        self.tableWidget.verticalHeader().setDefaultSectionSize(23)
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(490, 240, 111, 51))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.pushButton_2.setFont(font)
        icon2 = QtGui.QIcon()
        icon2.addPixmap(QtGui.QPixmap(":/icon/exit.png"), QtGui.QIcon.Normal, QtGui.QIcon.Off)
        self.pushButton_2.setIcon(icon2)
        self.pushButton_2.setIconSize(QtCore.QSize(25, 25))
        self.pushButton_2.setObjectName("pushButton_2")
        self.pushButton_2.clicked.connect(self.exit)
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(1030, 310, 181, 31))
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        MainWindow.setCentralWidget(self.centralwidget)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

    def retranslateUi(self, MainWindow):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "Vekon Excel Harcama Oluşturucu"))
        self.pushButton.setText(_translate("MainWindow", "EKLE"))
        self.label.setText(_translate("MainWindow", "FİRMA İSMİ"))
        self.pushButton_3.setText(_translate("MainWindow", "EXCEL"))
        self.label_3.setText(_translate("MainWindow", "    1          2        3     4     5        6          7       8        9"))
        self.label_7.setText(_translate("MainWindow", "MOBİL UYGULAMADAN DÖNÜŞTÜR"))
        self.pushButton_4.setText(_translate("MainWindow", "EXCEL"))
        self.label_4.setText(_translate("MainWindow", "MERT FINDIKLI"))
        item = self.tableWidget.horizontalHeaderItem(0)
        item.setText(_translate("MainWindow", "TARİH"))
        item = self.tableWidget.horizontalHeaderItem(1)
        item.setText(_translate("MainWindow", "FİŞ NUMASINI"))
        item = self.tableWidget.horizontalHeaderItem(2)
        item.setText(_translate("MainWindow", "FİRMA ADI"))
        item = self.tableWidget.horizontalHeaderItem(3)
        item.setText(_translate("MainWindow", "KİŞİ SAYISI"))
        item = self.tableWidget.horizontalHeaderItem(4)
        item.setText(_translate("MainWindow", "GİDER ÇEŞİTİ"))
        item = self.tableWidget.horizontalHeaderItem(5)
        item.setText(_translate("MainWindow", "FİYAT BİLGİSİ"))
        self.pushButton_2.setText(_translate("MainWindow", "ÇIKIŞ"))
        self.label_5.setText(_translate("MainWindow", "TOPLAM MİKTAR : 0"))
        self.toplam = 0
        self.fisid = 1
        self.fis = 1
        self.anatoplam = 0
        self.tarih = datetime.now()
        self.label_5.setText(_translate("MainWindow", "TOPLAM MİKTAR : 0"))
    def exit(self):
        exit()
    def ileri(self):
        if self.toplam == 0:
                cunsor.execute("Insert into ana (isadi) Values('{}')".format(self.lineEdit.text().upper()))
                self.label.setText("FİRMA ADI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......",1000)
        elif self.toplam == 1:
                cunsor.execute("update ana set firma = '{}'".format(self.lineEdit.text().upper()))
                self.label.setText("AVANSI ALAN ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 2:
                cunsor.execute("update  ana set avansialan = '{}'".format(self.lineEdit.text().upper()))
                self.label.setText("VERİLEN AVANS TUTARI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
        elif self.toplam == 3:
                cunsor.execute("update  ana set verilen = '{}'".format(self.lineEdit.text().upper()))
                self.label.setText("FİŞ TARİHİ ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
                sayi = str(self.fis) + ":" + "FİŞİ GİRİYORSUNUZ"
                self.label_6.setText(sayi)
        elif self.toplam == 4:
                cunsor.execute("Insert into FİRMA (id,tarih) Values('{}','{}')".format(self.fisid,
                                            self.lineEdit.text().upper()))
                self.label.setText("BELGE NO ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
                self.verisorgu()
                sayi = str(self.fis) + ":" + "FİŞİ GİRİYORSUNUZ"
                self.label_6.setText(sayi)
        elif self.toplam == 5:
                cunsor.execute("update  FİRMA set belge = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                        self.fisid))
                self.label.setText("HARCAMA FİRMA ADI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
                self.verisorgu()
        elif self.toplam == 6:
                cunsor.execute("update FİRMA set FİRMA = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                       self.fisid))
                self.label.setText("KİŞİ SAYISI ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
                self.verisorgu()
        elif self.toplam == 7:
                cunsor.execute("update FİRMA set Kişi = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                      self.fisid))
                self.label.setText("ÜSTEKİLERDEN HANGİSİ  ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
                self.verisorgu()
        elif self.toplam == 8:
                cunsor.execute("update FİRMA set gider = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                       self.fisid))
                self.label.setText("FİYAT NEDİR ")
                self.lineEdit.clear()
                self.toplam += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
                self.verisorgu()
        elif self.toplam == 9:
                cunsor.execute("update FİRMA set toplam = '{}' where id = '{}'".format(self.lineEdit.text().upper(),
                                                                                        self.fisid))
                self.label.setText("FİŞ TARİHİ  ")
                self.toplam -= 5
                self.fisid += 1
                self.statusbar.showMessage("EKLEME BAŞARILI......", 1000)
                self.verisorgu()
                self.fis +=1
                self.deger = int(self.lineEdit.text())
                self.anatoplam += self.deger
                self.label_5.setText("TOPLAM : "+str(self.anatoplam))
                self.lineEdit.clear()
                sayi = str(self.fis) + ":" + "FİŞİ GİRİYORSUNUZ"
                self.label_6.setText(sayi)
        baglanti.commit()
    def verisorgu(self):
        sorgu = "Select tarih,belge,FİRMA,Kişi,gider,toplam From FİRMA"
        cunsor.execute(sorgu)
        for k, ka in enumerate(cunsor):
            for a, ak in enumerate(ka):
                self.tableWidget.setItem(k,a,QtWidgets.QTableWidgetItem(str(ak)))
    def kaydet(self):
        import shutil
        shutil.copy("gerekli/test.xlsx", "./")
        self.dosya_ac = openpyxl.load_workbook("./test.xlsx")
        self.sayfa_ac = self.dosya_ac["Vekon Harcama"]
        self.sayfa_sayisi = self.sayfa_ac.max_row
        cunsor.execute("SELECT * FROM ana")
        self.islem = cunsor.fetchall()
        cunsor.execute("SELECT * FROM FİRMA")
        self.islem1 = cunsor.fetchall()
        self.toplam = 10
        self.toplamfiyat = 0
        for i in self.islem:
            self.sayfa_ac.cell(1, 3, value="{}".format(i[0]))
            self.sayfa_ac.cell(2, 3, value="{}".format(i[1]))
            self.sayfa_ac.cell(3, 3, value="{}".format(i[2]))
            self.sayfa_ac.cell(5, 3, value="{}".format(i[3]))
            self.sayfa_ac.cell(45, 2, value="{}".format(i[2]))
        for e in self.islem1:
            self.sayfa_ac.cell(self.toplam, 1, value="{}".format(e[1]))
            self.sayfa_ac.cell(self.toplam, 2, value="{}".format(e[2]))
            self.sayfa_ac.cell(self.toplam, 3, value="{}".format(e[3]))
            self.sayfa_ac.cell(self.toplam, 4, value="{}".format(e[4]))
            if e[5] == "1":
                self.sayfa_ac.cell(self.toplam, 5, value="{}".format("X"))
            elif e[5] == "2":
                self.sayfa_ac.cell(self.toplam, 6, value="{}".format("X"))
            elif e[5] == "3":
                self.sayfa_ac.cell(self.toplam, 7, value="{}".format("X"))
            elif e[5] == "4":
                self.sayfa_ac.cell(self.toplam, 8, value="{}".format("X"))
            elif e[5] == "5":
                self.sayfa_ac.cell(self.toplam, 9, value="{}".format("X"))
            elif e[5] == "6":
                self.sayfa_ac.cell(self.toplam, 10, value="{}".format("X"))
            elif e[5] == "7":
                self.sayfa_ac.cell(self.toplam, 11, value="{}".format("X"))
            elif e[5] == "8":
                self.sayfa_ac.cell(self.toplam, 12, value="{}".format("X"))
            elif e[5] == "9":
                self.sayfa_ac.cell(self.toplam, 13, value="{}".format("X"))
            self.sayfa_ac.cell(self.toplam, 16, value="{}".format(float(e[6])))
            self.toplam += 1
            self.toplamfiyat += e[6]
        self.sayfa_ac.cell(45, 1, value="{}".format(self.tarih.date()))
        self.sayfa_ac.cell(41, 16, value="{}".format(float(self.toplamfiyat)))
        self.dosya_ac.save("./test.xlsx")
        os.rename("test.xlsx", "{}_{}_{}_HARCAMA_BELGESİ_{}.xlsx".format((str(i[2]).strip("\n\n")),(str(i[0]).strip("\n\n")),(str(i[1]).strip("\n\n")),self.tarih.date()))
        cunsor.execute("delete from FİRMA")
        cunsor.execute("delete from ana")
        baglanti.commit()
    def mobil(self):
        # ANA ELEMETLER.
        self.kod = self.lineEdit_2.text()
        print(self.kod)
        if len(self.kod) == 0:
            self.statusbar.showMessage("LÜTFEN KODU GİRİNİZ......", 1000)
        self.isadi = []
        self.firmaadi = []
        self.avansialan = []
        self.verilenavans = []
        self.mail = []
        cred = credentials.Certificate('vkharcama-firebase-adminsdk-nba1g-0dcf5122a7.json')
        app = firebase_admin.initialize_app(cred)
        db = firestore.client()
        users_ref = db.collection(u'{}'.format(self.kod))
        docs = users_ref.stream()
        self.list = []
        self.fis = {}
        self.fisler = 0
        for doc in docs:
            self.mert = f'{doc.to_dict()}'
            self.mert1 = str(self.mert).strip("{ }")
            self.mert2 = str(self.mert1).split(",")
            for i in self.mert2:
                self.list.append(str(i).replace("'", '').replace(' ', '').split(":"))
            for i in self.list:
                if (i[0] == "firmaadi"):
                    self.firmaadi.append(i[1])
                if (i[0] == "avansialan"):
                    self.avansialan.append(i[1])
                if (i[0] == "mail"):
                    self.mail.append(i[1])
                if (i[0] == "verilenavans"):
                    self.verilenavans.append(i[1])
                if (i[0] == "isnumarasi"):
                    self.isadi.append(i[1])
                if ('fistarih' in i[0]):
                    self.fisler += 1
        self.fissayisi = 0
        self.karakter = 1
        while (self.fissayisi <= self.fisler):
            for i in self.list:
                if (i[0] == "fistarih{}".format(self.karakter)):
                    self.fis["fistarihi{}".format(self.karakter)] = i[1].upper()
                if (i[0] == "fisbelgeno{}".format(self.karakter)):
                    self.fis["belgeno{}".format(self.karakter)] = i[1].upper()
                if (i[0] == "fisfirmaadi{}".format(self.karakter)):
                    self.fis["firmaadi{}".format(self.karakter)] = i[1].upper()
                if (i[0] == "fiskisisayisi{}".format(self.karakter)):
                    self.fis["kisi{}".format(self.karakter)] = i[1].upper()
                if (i[0] == "fisgidercesiti{}".format(self.karakter)):
                    self.fis["cesit{}".format(self.karakter)] = i[1].upper()
                if (i[0] == "fisfiyat{}".format(self.karakter)):
                    self.fis["fiyat{}".format(self.karakter)] = i[1].upper()
            self.karakter += 1
            self.fissayisi += 1
        self.excelvericek()

    def excelvericek(self):
        import shutil
        shutil.copy("gerekli/test.xlsx", "./")
        self.dosya_ac = openpyxl.load_workbook("./test.xlsx")
        self.sayfa_ac = self.dosya_ac["Vekon Harcama"]
        self.sayfa_sayisi = self.sayfa_ac.max_row
        self.toplam = 10
        self.toplamfiyat = 0
        # ANA ELEMENT
        print(self.isadi)
        self.isadi1 = str(self.isadi[0]).upper()
        self.firmaadi1 = str(self.firmaadi[0]).upper()
        self.avansialan1 = str(self.avansialan[0]).upper()
        self.verilenavans1 = str(self.verilenavans[0]).upper()

        self.sayfa_ac.cell(1, 3, value="{}".format(self.isadi1))
        self.sayfa_ac.cell(2, 3, value="{}".format(self.firmaadi1))
        self.sayfa_ac.cell(3, 3, value="{}".format(self.avansialan1))
        self.sayfa_ac.cell(5, 3, value="{}".format(self.verilenavans1))
        self.sayfa_ac.cell(45, 2, value="{}".format(self.avansialan1))
        self.fisci =1
        self.sayi =1
        while(self.fisci <= self.fisler):
            self.sayfa_ac.cell(self.toplam, 1, value=self.fis['fistarihi{}'.format(self.sayi)])
            self.sayfa_ac.cell(self.toplam, 2, value=self.fis['belgeno{}'.format(self.sayi)])
            self.sayfa_ac.cell(self.toplam, 3, value=self.fis['firmaadi{}'.format(self.sayi)])
            self.sayfa_ac.cell(self.toplam, 4, value=self.fis['kisi{}'.format(self.sayi)])
            if self.fis['cesit{}'.format(self.sayi)] == "MALZEME":
                self.sayfa_ac.cell(self.toplam, 5, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "AKARYAKIT":
                self.sayfa_ac.cell(self.toplam, 6, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "YOL":
                self.sayfa_ac.cell(self.toplam, 7, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "OTEL":
                self.sayfa_ac.cell(self.toplam, 8, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "YEMEK":
                self.sayfa_ac.cell(self.toplam, 9, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "HABERLEŞME":
                self.sayfa_ac.cell(self.toplam, 10, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "SAĞLIK":
                self.sayfa_ac.cell(self.toplam, 11, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "SAİR":
                self.sayfa_ac.cell(self.toplam, 12, value="{}".format("X"))
            elif self.fis['cesit{}'.format(self.sayi)] == "BELGESİZ":
                self.sayfa_ac.cell(self.toplam, 13, value="{}".format("X"))
            self.sayfa_ac.cell(self.toplam, 16, value=self.fis['fiyat{}'.format(self.sayi)])
            self.toplam += 1
            self.toplamfiyat += float(self.fis['fiyat{}'.format(self.sayi)])
            self.fisci +=1
            self.sayi +=1
        self.sayfa_ac.cell(45, 1, value="{}".format(self.tarih.date()))
        self.sayfa_ac.cell(41, 16, value="{}".format(float(self.toplamfiyat)))
        self.dosya_ac.save("./test.xlsx")
        os.rename("test.xlsx",
                  "{}_{}_{}_HARCAMA_BELGESİ_{}.xlsx".format(self.avansialan1,
                                                            self.isadi1,self.firmaadi1,self.tarih.date()))
import resim_rc
if __name__ == "__main__":
        import sys
        app = QtWidgets.QApplication(sys.argv)
        pencere = QtWidgets.QMainWindow()
        ui = Ui_MainWindow()
        ui.setupUi(pencere)
        pencere.show()
        sys.exit(app.exec_())